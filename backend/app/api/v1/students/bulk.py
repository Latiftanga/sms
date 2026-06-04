import io
import uuid
from datetime import date

from fastapi import APIRouter, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select

from app.api.deps import CurrentUser, SessionDep, require
from app.core.permissions import Permission
from app.models.academic import AcademicYear, Class
from app.models.school import School, SchoolConfig
from app.models.student import Guardian, Student, StudentClassEnrollment, StudentTermEnrollment
from app.models.academic import AcademicTerm
from app.schemas.staff import BulkRowError, BulkUploadResponse

router = APIRouter()

_TEMPLATE_COLUMNS = [
    # (field_key, header_label, col_width)
    ("first_name",          "First Name *",       18),
    ("middle_name",         "Middle Name",         18),
    ("last_name",           "Last Name *",         18),
    ("gender",              "Gender *",            12),
    ("date_of_birth",       "Date of Birth",       14),
    ("place_of_birth",      "Place of Birth",      20),
    ("nationality",         "Nationality",         16),
    ("religion",            "Religion",            16),
    ("admission_date",      "Admission Date",      14),
    ("admission_number",    "Admission Number",    18),
    ("previous_school",     "Previous School",     24),
    ("class_name",          "Class Name",          16),
    ("student_type",        "Student Type",        14),
    ("guardian_first_name", "Guardian First Name", 20),
    ("guardian_last_name",  "Guardian Last Name",  20),
    ("guardian_relationship","Guardian Relationship",22),
    ("guardian_phone",      "Guardian Phone",      16),
]

_DROPDOWNS: dict[str, list[str]] = {
    "gender":                ["MALE", "FEMALE"],
    "student_type":          ["DAY", "BOARDING"],
    "guardian_relationship": ["FATHER", "MOTHER", "GUARDIAN"],
}

_EXAMPLE_ROWS = [
    {
        "first_name": "Kwame", "middle_name": "Asante", "last_name": "Mensah",
        "gender": "MALE", "date_of_birth": "2012-04-15",
        "nationality": "Ghanaian", "admission_date": "2023-09-04",
        "class_name": "Basic 4A", "student_type": "DAY",
        "guardian_first_name": "Kofi", "guardian_last_name": "Mensah",
        "guardian_relationship": "FATHER", "guardian_phone": "0244123456",
    },
    {
        "first_name": "Abena", "last_name": "Owusu",
        "gender": "FEMALE", "date_of_birth": "2011-08-22",
        "nationality": "Ghanaian",
        "class_name": "Basic 5A", "student_type": "DAY",
        "guardian_first_name": "Ama", "guardian_last_name": "Owusu",
        "guardian_relationship": "MOTHER", "guardian_phone": "0201987654",
    },
]

_LABEL_TO_FIELD: dict[str, str] = {
    label.lower().rstrip(" *"): field
    for field, label, _ in _TEMPLATE_COLUMNS
}

_DATE_FIELDS = {"date_of_birth", "admission_date"}
_REQUIRED = {"first_name", "last_name", "gender"}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _hex(color: str) -> str:
    return color.lstrip("#").upper()

def _darken(color: str, f: float = 0.55) -> str:
    h = _hex(color); r, g, b = int(h[0:2],16), int(h[2:4],16), int(h[4:6],16)
    return f"{int(r*f):02X}{int(g*f):02X}{int(b*f):02X}"

def _tint(color: str, s: float = 0.08) -> str:
    h = _hex(color); r, g, b = int(h[0:2],16), int(h[2:4],16), int(h[4:6],16)
    return f"{int(r*s+255*(1-s)):02X}{int(g*s+255*(1-s)):02X}{int(b*s+255*(1-s)):02X}"


async def _generate_register_number(school: School, year: AcademicYear, session) -> str:
    config = await session.scalar(select(SchoolConfig).where(SchoolConfig.school_id == school.id))
    pattern = (config.register_number_pattern if config else None) or "{code}/{year}/{seq:04d}"
    seq = await session.scalar(
        select(func.count(StudentClassEnrollment.id))
        .where(StudentClassEnrollment.academic_year_id == year.id)
    ) or 0
    return pattern.format(code=school.code, year=year.name[:4], seq=seq + 1)


# ── Template download ─────────────────────────────────────────────────────────

@router.get("/bulk/template", dependencies=[require(Permission.ENROLL_STUDENTS)])
async def download_template(user: CurrentUser, session: SessionDep):
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    school = await session.scalar(select(School).where(School.id == user.school_id))
    school_name = school.name if school else "School"
    accent = school.accent_color if school else "#185FA5"

    c_accent  = _hex(accent)
    c_dark    = _darken(accent)
    c_tint    = _tint(accent, 0.08)
    c_example = _tint(accent, 0.12)
    c_border  = _tint(accent, 0.30)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Import"

    total_cols = len(_TEMPLATE_COLUMNS)
    thin = Side(style="thin", color=c_border)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    tc = ws.cell(row=1, column=1, value=f"{school_name} — Student Import Template")
    tc.font = Font(bold=True, size=13, color="FFFFFF")
    tc.fill = PatternFill("solid", fgColor=c_dark)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Sub-title
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    sc = ws.cell(row=2, column=1,
        value="Required: first_name, last_name, gender  ·  Dates: YYYY-MM-DD  ·  class_name must match exactly  ·  Delete example rows before importing")
    sc.font = Font(italic=True, size=9, color="444444")
    sc.fill = PatternFill("solid", fgColor=c_tint)
    sc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Header row
    for col_idx, (field, label, col_width) in enumerate(_TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=label)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=c_accent)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[3].height = 30

    # Example rows
    fields = [f for f, _, _ in _TEMPLATE_COLUMNS]
    for row_off, example in enumerate(_EXAMPLE_ROWS, start=4):
        for col_idx, field in enumerate(fields, start=1):
            cell = ws.cell(row=row_off, column=col_idx, value=example.get(field, ""))
            cell.fill = PatternFill("solid", fgColor=c_example)
            cell.font = Font(size=10, color="333333")
            cell.alignment = Alignment(vertical="center")
            cell.border = border
        ws.row_dimensions[row_off].height = 18

    # Data rows
    for row in range(6, 506):
        for col_idx in range(1, total_cols + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = Font(size=10)
            cell.alignment = Alignment(vertical="center")
            cell.border = border
        ws.row_dimensions[row].height = 18

    # Dropdowns
    for col_idx, (field, _, _) in enumerate(_TEMPLATE_COLUMNS, start=1):
        if field not in _DROPDOWNS:
            continue
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(_DROPDOWNS[field]) + '"',
            allow_blank=True, showDropDown=False,
            showErrorMessage=True,
            errorTitle="Invalid value",
            error=f'Use the dropdown for {field.replace("_", " ")}.',
        )
        dv.sqref = f"{col_letter}4:{col_letter}505"
        ws.add_data_validation(dv)

    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe = school_name.replace(" ", "_").replace("/", "-")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe}_Student_Import.xlsx"'},
    )


# ── Parsing ───────────────────────────────────────────────────────────────────

def _parse_excel(content: bytes) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    known = set(_LABEL_TO_FIELD) | {f for f, _, _ in _TEMPLATE_COLUMNS}
    header_idx = next(
        (i for i, row in enumerate(rows)
         if row[0] is not None and str(row[0]).strip().lower().rstrip(" *") in known),
        0
    )
    raw_headers = [str(h).strip().lower().rstrip(" *") if h is not None else "" for h in rows[header_idx]]
    headers = [_LABEL_TO_FIELD.get(h, h) for h in raw_headers]
    return [dict(zip(headers, row)) for row in rows[header_idx + 1:]]


def _parse_csv(content: bytes) -> list[dict]:
    import csv
    text = content.decode("utf-8-sig")
    lines = [l for l in text.splitlines() if l.strip() and not l.strip().startswith("#")]
    return list(csv.DictReader(io.StringIO("\n".join(lines))))


# ── Upload ────────────────────────────────────────────────────────────────────

@router.post("/bulk", response_model=BulkUploadResponse, status_code=201,
             dependencies=[require(Permission.ENROLL_STUDENTS)])
async def bulk_upload(
    user: CurrentUser,
    session: SessionDep,
    file: UploadFile = File(...),
):
    content = await file.read()
    fname = (file.filename or "").lower()

    if fname.endswith(".csv"):
        rows = _parse_csv(content)
    elif fname.endswith((".xlsx", ".xls")):
        rows = _parse_excel(content)
    else:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Upload a .csv or .xlsx file")

    school_id = user.school_id
    school = await session.get(School, school_id)

    # Current academic year (needed for enrollment)
    current_year = await session.scalar(
        select(AcademicYear).where(
            AcademicYear.school_id == school_id,
            AcademicYear.is_current.is_(True),
        )
    )

    # Build class name → id map for fast lookup
    class_map: dict[str, uuid.UUID] = {}
    if current_year:
        class_rows = await session.scalars(
            select(Class).where(Class.school_id == school_id, Class.is_active.is_(True))
        )
        for c in class_rows:
            class_map[c.name.lower()] = c.id

    # Load all terms for current year (for auto term enrollment)
    year_terms: list = []
    if current_year:
        year_terms = list(await session.scalars(
            select(AcademicTerm).where(AcademicTerm.academic_year_id == current_year.id)
        ))

    created = 0
    skipped = 0
    errors: list[BulkRowError] = []

    for i, raw in enumerate(rows, start=2):
        row = {
            k.strip().lower() if k else "": (str(v).strip() if v is not None else "")
            for k, v in raw.items()
        }

        if not any(row.values()):
            skipped += 1
            continue

        # Required field validation
        missing = [f for f in _REQUIRED if not row.get(f)]
        if missing:
            errors.append(BulkRowError(row=i, field=missing[0], message=f"Required: {missing[0]}"))
            continue

        # Build student kwargs
        kwargs: dict = {
            "id": uuid.uuid4(),
            "school_id": school_id,
            "first_name": row["first_name"],
            "last_name": row["last_name"],
            "gender": row["gender"].upper() if row.get("gender") else "MALE",
            "nationality": row.get("nationality") or "Ghanaian",
            "is_active": True,
        }

        for field in ("middle_name", "place_of_birth", "religion",
                      "admission_number", "previous_school"):
            if row.get(field):
                kwargs[field] = row[field]

        # Date fields
        date_ok = True
        for field in _DATE_FIELDS:
            val = row.get(field)
            if val:
                try:
                    kwargs[field] = date.fromisoformat(val)
                except ValueError:
                    errors.append(BulkRowError(row=i, field=field, message=f"Invalid date '{val}' — use YYYY-MM-DD"))
                    date_ok = False
                    break
        if not date_ok:
            continue

        try:
            student = Student(**kwargs)
            session.add(student)
            await session.flush()

            # Guardian (if provided)
            g_first = row.get("guardian_first_name", "").strip()
            g_last = row.get("guardian_last_name", "").strip()
            if g_first and g_last:
                session.add(Guardian(
                    id=uuid.uuid4(),
                    student_id=student.id,
                    first_name=g_first,
                    last_name=g_last,
                    relationship_type=(row.get("guardian_relationship") or "GUARDIAN").upper(),
                    phone=row.get("guardian_phone") or None,
                    is_primary_contact=True,
                ))

            # Class enrollment (if class_name provided and current year exists)
            class_name_raw = row.get("class_name", "").strip().lower()
            if class_name_raw and current_year and class_name_raw in class_map:
                reg_num = await _generate_register_number(school, current_year, session)
                enrollment = StudentClassEnrollment(
                    id=uuid.uuid4(),
                    student_id=student.id,
                    class_id=class_map[class_name_raw],
                    academic_year_id=current_year.id,
                    student_type=(row.get("student_type") or "DAY").upper(),
                    register_number=reg_num,
                    status="ACTIVE",
                )
                session.add(enrollment)
                await session.flush()

                for term in year_terms:
                    session.add(StudentTermEnrollment(
                        id=uuid.uuid4(),
                        student_class_enrollment_id=enrollment.id,
                        academic_term_id=term.id,
                        enrolled_date=date.today(),
                        fee_status="NOT_APPLICABLE",
                    ))
            elif class_name_raw and class_name_raw not in class_map:
                # Non-fatal: student is created but not enrolled
                errors.append(BulkRowError(
                    row=i, field="class_name",
                    message=f"Class '{row.get('class_name')}' not found — student created without enrollment",
                ))

            await session.flush()
            created += 1

        except Exception as exc:
            errors.append(BulkRowError(row=i, field=None, message=str(exc)))

    if errors and created == 0:
        await session.rollback()
    else:
        await session.commit()

    return BulkUploadResponse(created=created, skipped=skipped, errors=errors)
