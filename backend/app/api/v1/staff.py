"""
Staff management API.

Routes:
  GET    /staff                            — paginated list
  POST   /staff                            — create staff member
  GET    /staff/{id}                       — detail with qualifications + promotions
  PATCH  /staff/{id}                       — update profile
  DELETE /staff/{id}                       — deactivate (soft delete)
  POST   /staff/{id}/account               — create login account (temp password)
  POST   /staff/{id}/photo                 — upload profile photo
  POST   /staff/{id}/qualifications        — add qualification
  PATCH  /staff/{id}/qualifications/{qid} — update qualification
  DELETE /staff/{id}/qualifications/{qid} — remove qualification
  GET    /staff/{id}/promotions            — rank history
  POST   /staff/{id}/promotions            — record new rank
  PATCH  /staff/{id}/promotions/{pid}      — update rank entry
  DELETE /staff/{id}/promotions/{pid}      — remove rank entry
  GET    /staff/{id}/permissions           — resolved permissions + overrides
  POST   /staff/{id}/permissions           — grant/revoke permission override
  DELETE /staff/{id}/permissions/{key}     — remove override
  PATCH  /staff/{id}/position              — change assigned position
  POST   /staff/bulk                       — bulk CSV/Excel upload
"""
import io
import secrets
from datetime import UTC, date, datetime, timedelta
from uuid import UUID

import httpx
from fastapi import APIRouter, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import exists, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.api.deps import CurrentUser, RedisDep, SessionDep, require
from app.core.config import settings
from app.core.permissions import ALL_PERMISSIONS, Permission
from app.core.security import hash_password
from app.models.staff import (
    StaffMember,
    StaffPermission,
    StaffPosition,
    StaffPromotion,
    StaffQualification,
)
from app.models.user import User, UserRole
from app.schemas.staff import (
    AccountCreateRequest,
    InviteResponse,
    BulkRowError,
    BulkUploadResponse,
    PermissionOverrideCreate,
    PermissionOverrideResponse,
    PromotionCreate,
    PromotionResponse,
    PromotionUpdate,
    QualificationCreate,
    QualificationResponse,
    QualificationUpdate,
    RoleAssignRequest,
    RoleResponse,
    StaffMemberCreate,
    StaffMemberDetail,
    StaffMemberResponse,
    StaffMemberUpdate,
    StaffPermissionsResponse,
    UserRoleResponse,
)
from app.schemas.common import PagedResponse
from app.services.permissions import invalidate_cache, resolve_all_permissions
from app.services.storage import get_storage

router = APIRouter(prefix="/staff", tags=["staff"])

ALLOWED_PHOTO_TYPES = {"image/jpeg", "image/png", "image/webp"}


# ── Helpers ───────────────────────────────────────────────────────────────────

# Designation → default system role codes assigned at account creation
_DESIGNATION_TO_ROLE_CODES: dict[str, list[str]] = {
    "HEADTEACHER":         ["HEADTEACHER"],
    "ASSISTANT_HEAD":      ["ASSISTANT_HEAD"],
    "BURSAR":              ["BURSAR"],
    "HOUSEMASTER":         ["HOUSEMASTER"],
    "SENIOR_HOUSEMASTER":  ["SENIOR_HOUSEMASTER"],
}


def _default_role_codes(member: StaffMember) -> list[str]:
    codes = _DESIGNATION_TO_ROLE_CODES.get(member.designation or "")
    if codes:
        return codes
    if member.category == "TEACHING":
        return ["CLASS_TEACHER"]
    return []


async def _assign_roles_by_codes(
    user: User, codes: list[str], assigned_by_id: UUID, session: AsyncSession
) -> None:
    if not codes:
        return
    positions = list(await session.scalars(
        select(StaffPosition).where(
            StaffPosition.code.in_(codes),
            StaffPosition.is_system_template.is_(True),
        )
    ))
    now = datetime.now(UTC)
    for pos in positions:
        exists_already = await session.scalar(
            select(UserRole).where(
                UserRole.user_id == user.id,
                UserRole.role_id == pos.id,
            )
        )
        if not exists_already:
            session.add(UserRole(
                user_id=user.id,
                role_id=pos.id,
                assigned_by=assigned_by_id,
                assigned_at=now,
            ))


async def _get_member(staff_id: UUID, school_id: UUID, session) -> StaffMember:
    member = await session.scalar(
        select(StaffMember).where(
            StaffMember.id == staff_id,
            StaffMember.school_id == school_id,
        )
    )
    if not member:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Staff member not found")
    return member


async def _linked_user(member: StaffMember, session) -> User | None:
    return await session.scalar(
        select(User).where(User.staff_member_id == member.id)
    )


def _current_rank(promotions: list[StaffPromotion]) -> str | None:
    if not promotions:
        return None
    return max(promotions, key=lambda p: p.date_promoted).rank


def _to_response(member: StaffMember, has_account: bool = False) -> StaffMemberResponse:
    # Only read promotions if already eagerly loaded — never trigger a lazy load in async context
    loaded_promotions = member.__dict__.get("promotions")
    return StaffMemberResponse.model_validate({
        **member.__dict__,
        "current_rank": _current_rank(loaded_promotions) if loaded_promotions else None,
        "has_account": has_account,
    })





# ── List ──────────────────────────────────────────────────────────────────────

@router.get("", response_model=PagedResponse[StaffMemberResponse],
            dependencies=[require(Permission.VIEW_STAFF)])
async def list_staff(
    user: CurrentUser,
    session: SessionDep,
    skip: int = Query(0, ge=0),
    limit: int = Query(25, ge=1, le=100),
    category: str | None = None,
    is_active: bool | None = None,
    search: str | None = None,
):
    school_id = user.school_id

    # Build reusable filter conditions
    conditions = [StaffMember.school_id == school_id]
    if category:
        conditions.append(StaffMember.category == category.upper())
    if is_active is not None:
        conditions.append(StaffMember.is_active == is_active)
    if search:
        term = f"%{search}%"
        conditions.append(
            StaffMember.first_name.ilike(term)
            | StaffMember.last_name.ilike(term)
            | StaffMember.staff_id.ilike(term)
            | StaffMember.ges_staff_id.ilike(term)
        )

    total = await session.scalar(
        select(func.count(StaffMember.id)).where(*conditions)
    )

    # Correlated subquery for current rank — avoids loading full promotion
    # history for every row via selectinload
    rank_sq = (
        select(StaffPromotion.rank)
        .where(StaffPromotion.staff_member_id == StaffMember.id)
        .order_by(StaffPromotion.date_promoted.desc())
        .limit(1)
        .correlate(StaffMember)
        .scalar_subquery()
    )

    rows = list(await session.execute(
        select(StaffMember, rank_sq.label("current_rank"))
        .where(*conditions)
        .order_by(StaffMember.last_name, StaffMember.first_name)
        .offset(skip)
        .limit(limit)
    ))

    # Batch-check which members have login accounts
    ids = [r[0].id for r in rows]
    linked_ids: set[UUID] = set()
    if ids:
        linked_ids = set(await session.scalars(
            select(User.staff_member_id).where(User.staff_member_id.in_(ids))
        ))

    items = [
        StaffMemberResponse.model_validate({
            **r[0].__dict__,
            "current_rank": r[1],
            "has_account": r[0].id in linked_ids,
        })
        for r in rows
    ]
    return PagedResponse(items=items, total=total or 0, skip=skip, limit=limit)


# ── Create ────────────────────────────────────────────────────────────────────

@router.post("", response_model=StaffMemberResponse, status_code=201,
             dependencies=[require(Permission.MANAGE_STAFF)])
async def create_staff(
    body: StaffMemberCreate,
    user: CurrentUser,
    session: SessionDep,
):
    member = StaffMember(school_id=user.school_id, **body.model_dump())
    session.add(member)
    await session.commit()
    await session.refresh(member)
    return _to_response(member)


# ── Detail ────────────────────────────────────────────────────────────────────

@router.get("/{staff_id}", response_model=StaffMemberDetail,
            dependencies=[require(Permission.VIEW_STAFF)])
async def get_staff(staff_id: UUID, user: CurrentUser, session: SessionDep):

    has_acct_sq = exists().where(User.staff_member_id == staff_id)
    invite_pending_sq = exists().where(
        User.staff_member_id == staff_id,
        User.invite_token.isnot(None),
        User.is_active.is_(False),
    )

    result = await session.execute(
        select(StaffMember, has_acct_sq, invite_pending_sq)
        .where(
            StaffMember.id == staff_id,
            StaffMember.school_id == user.school_id,
        )
        .options(
            selectinload(StaffMember.qualifications),
            selectinload(StaffMember.promotions),
        )
    )
    row = result.one_or_none()
    if not row:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Staff member not found")

    member, has_account, invite_pending = row
    detail = StaffMemberDetail.model_validate({
        **member.__dict__,
        "current_rank": _current_rank(member.promotions),
        "has_account": has_account,
        "invite_pending": invite_pending,
        "qualifications": [QualificationResponse.model_validate(q) for q in member.qualifications],
        "promotions": sorted(
            [PromotionResponse.model_validate(p) for p in member.promotions],
            key=lambda p: p.date_promoted,
            reverse=True,
        ),
    })
    return detail


# ── Update ────────────────────────────────────────────────────────────────────

@router.patch("/{staff_id}", response_model=StaffMemberResponse,
              dependencies=[require(Permission.MANAGE_STAFF)])
async def update_staff(
    staff_id: UUID,
    body: StaffMemberUpdate,
    user: CurrentUser,
    session: SessionDep,
):
    member = await _get_member(staff_id, user.school_id, session)
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(member, field, value)
    await session.commit()
    await session.refresh(member)
    return _to_response(member)


# ── Soft-delete ───────────────────────────────────────────────────────────────

@router.delete("/{staff_id}", status_code=204,
               dependencies=[require(Permission.MANAGE_STAFF)])
async def deactivate_staff(
    staff_id: UUID,
    user: CurrentUser,
    session: SessionDep,
):
    member = await _get_member(staff_id, user.school_id, session)
    member.is_active = False
    # Deactivate linked user account too
    linked = await _linked_user(member, session)
    if linked:
        linked.is_active = False
    await session.commit()


# ── Photo upload ──────────────────────────────────────────────────────────────

@router.post("/{staff_id}/photo", response_model=StaffMemberResponse,
             dependencies=[require(Permission.MANAGE_STAFF)])
async def upload_photo(
    staff_id: UUID,
    user: CurrentUser,
    session: SessionDep,
    file: UploadFile = File(...),
):
    if file.content_type not in ALLOWED_PHOTO_TYPES:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "JPEG, PNG or WebP only")

    member = await _get_member(staff_id, user.school_id, session)
    storage = get_storage()

    if member.photo_url:
        await storage.delete(member.photo_url)

    member.photo_url = await storage.upload(file, folder="staff-photos")
    await session.commit()
    await session.refresh(member)
    return _to_response(member)


# ── Invite ────────────────────────────────────────────────────────────────────

async def _send_invite_sms(phone: str, invite_token: str) -> None:
    invite_url = f"{settings.FRONTEND_URL}/invite/{invite_token}"
    msg = f"You've been invited to TTEK SIS. Set up your account: {invite_url}"
    async with httpx.AsyncClient(timeout=8) as client:
        resp = await client.post(
            "https://api.africastalking.com/version1/messaging",
            headers={"apiKey": settings.AT_API_KEY or "", "Accept": "application/json"},
            data={"username": settings.AT_USERNAME, "to": phone, "message": msg, "from": settings.AT_SENDER_ID},
        )
        resp.raise_for_status()


@router.post("/{staff_id}/invite", response_model=InviteResponse, status_code=201,
             dependencies=[require(Permission.MANAGE_USERS)])
async def send_invite(
    staff_id: UUID,
    body: AccountCreateRequest,
    user: CurrentUser,
    session: SessionDep,
):
    member = await _get_member(staff_id, user.school_id, session)
    existing = await _linked_user(member, session)

    if existing and existing.is_active:
        raise HTTPException(status.HTTP_409_CONFLICT, "Staff member already has an active account")

    # Ensure email isn't taken by a different user
    email_owner = await session.scalar(select(User).where(User.email == body.email))
    if email_owner and (existing is None or email_owner.id != existing.id):
        raise HTTPException(status.HTTP_409_CONFLICT, "Email already in use")

    token = secrets.token_urlsafe(32)
    expires_at = datetime.now(UTC) + timedelta(days=7)

    if existing:
        # Resend: refresh token on the pending account
        existing.email = str(body.email)
        existing.invite_token = token
        existing.invite_expires_at = expires_at
    else:
        new_user = User(
            email=str(body.email),
            password_hash=hash_password(secrets.token_hex(32)),  # unreachable placeholder
            system_role="SCHOOL_STAFF",
            school_id=user.school_id,
            staff_member_id=member.id,
            is_active=False,
            is_verified=False,
            must_change_password=False,
            invite_token=token,
            invite_expires_at=expires_at,
        )
        session.add(new_user)
        await session.flush()
        codes = _default_role_codes(member)
        await _assign_roles_by_codes(new_user, codes, user.id, session)

    await session.commit()

    sms_sent = False
    if member.phone and settings.AT_API_KEY:
        try:
            await _send_invite_sms(member.phone, token)
            sms_sent = True
        except Exception:
            pass

    return InviteResponse(invite_token=token, email=str(body.email), sms_sent=sms_sent)


# ── Admin password reset — sends a new invite link ───────────────────────────

@router.post("/{staff_id}/reset-password", response_model=InviteResponse,
             dependencies=[require(Permission.MANAGE_USERS)])
async def reset_password(
    staff_id: UUID,
    user: CurrentUser,
    session: SessionDep,
):
    member = await _get_member(staff_id, user.school_id, session)
    linked = await _linked_user(member, session)
    if not linked:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "No account found for this staff member")

    token = secrets.token_urlsafe(32)
    expires_at = datetime.now(UTC) + timedelta(days=7)
    linked.is_active = False
    linked.is_verified = False
    linked.invite_token = token
    linked.invite_expires_at = expires_at
    await session.commit()

    sms_sent = False
    if member.phone and settings.AT_API_KEY:
        try:
            await _send_invite_sms(member.phone, token)
            sms_sent = True
        except Exception:
            pass

    return InviteResponse(invite_token=token, email=linked.email, sms_sent=sms_sent)


# ── Reactivate staff ──────────────────────────────────────────────────────────

@router.post("/{staff_id}/reactivate", status_code=204,
             dependencies=[require(Permission.MANAGE_STAFF)])
async def reactivate_staff(
    staff_id: UUID,
    user: CurrentUser,
    session: SessionDep,
):
    member = await _get_member(staff_id, user.school_id, session)
    member.is_active = True
    linked = await _linked_user(member, session)
    if linked:
        linked.is_active = True
    await session.commit()


# ── Qualifications ────────────────────────────────────────────────────────────

@router.post("/{staff_id}/qualifications", response_model=QualificationResponse, status_code=201,
             dependencies=[require(Permission.MANAGE_STAFF)])
async def add_qualification(
    staff_id: UUID,
    body: QualificationCreate,
    user: CurrentUser,
    session: SessionDep,
):
    member = await _get_member(staff_id, user.school_id, session)
    q = StaffQualification(staff_member_id=member.id, **body.model_dump())
    session.add(q)
    await session.commit()
    await session.refresh(q)
    return QualificationResponse.model_validate(q)


@router.patch("/{staff_id}/qualifications/{qual_id}", response_model=QualificationResponse,
              dependencies=[require(Permission.MANAGE_STAFF)])
async def update_qualification(
    staff_id: UUID,
    qual_id: UUID,
    body: QualificationUpdate,
    user: CurrentUser,
    session: SessionDep,
):
    member = await _get_member(staff_id, user.school_id, session)
    qual = await session.scalar(
        select(StaffQualification).where(
            StaffQualification.id == qual_id,
            StaffQualification.staff_member_id == member.id,
        )
    )
    if not qual:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Qualification not found")
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(qual, field, value)
    await session.commit()
    await session.refresh(qual)
    return QualificationResponse.model_validate(qual)


@router.delete("/{staff_id}/qualifications/{qual_id}", status_code=204,
               dependencies=[require(Permission.MANAGE_STAFF)])
async def remove_qualification(
    staff_id: UUID,
    qual_id: UUID,
    user: CurrentUser,
    session: SessionDep,
):
    member = await _get_member(staff_id, user.school_id, session)
    qual = await session.scalar(
        select(StaffQualification).where(
            StaffQualification.id == qual_id,
            StaffQualification.staff_member_id == member.id,
        )
    )
    if not qual:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Qualification not found")
    await session.delete(qual)
    await session.commit()


# ── Promotions / GES rank history ─────────────────────────────────────────────

@router.get("/{staff_id}/promotions", response_model=list[PromotionResponse],
            dependencies=[require(Permission.VIEW_STAFF)])
async def list_promotions(
    staff_id: UUID,
    user: CurrentUser,
    session: SessionDep,
):
    member = await _get_member(staff_id, user.school_id, session)
    rows = list(await session.scalars(
        select(StaffPromotion)
        .where(StaffPromotion.staff_member_id == member.id)
        .order_by(StaffPromotion.date_promoted.desc())
    ))
    return [PromotionResponse.model_validate(r) for r in rows]


@router.post("/{staff_id}/promotions", response_model=PromotionResponse, status_code=201,
             dependencies=[require(Permission.MANAGE_PROMOTIONS)])
async def record_promotion(
    staff_id: UUID,
    body: PromotionCreate,
    user: CurrentUser,
    session: SessionDep,
):
    member = await _get_member(staff_id, user.school_id, session)
    row = StaffPromotion(
        staff_member_id=member.id,
        rank=body.rank,
        date_promoted=body.date_promoted,
        date_recorded=date.today(),
        recorded_by=user.id,
    )
    session.add(row)
    await session.commit()
    await session.refresh(row)
    return PromotionResponse.model_validate(row)


@router.patch("/{staff_id}/promotions/{prom_id}", response_model=PromotionResponse,
              dependencies=[require(Permission.MANAGE_PROMOTIONS)])
async def update_promotion(
    staff_id: UUID,
    prom_id: UUID,
    body: PromotionUpdate,
    user: CurrentUser,
    session: SessionDep,
):
    member = await _get_member(staff_id, user.school_id, session)
    row = await session.scalar(
        select(StaffPromotion).where(
            StaffPromotion.id == prom_id,
            StaffPromotion.staff_member_id == member.id,
        )
    )
    if not row:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Promotion record not found")
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(row, field, value)
    await session.commit()
    await session.refresh(row)
    return PromotionResponse.model_validate(row)


@router.delete("/{staff_id}/promotions/{prom_id}", status_code=204,
               dependencies=[require(Permission.MANAGE_PROMOTIONS)])
async def remove_promotion(
    staff_id: UUID,
    prom_id: UUID,
    user: CurrentUser,
    session: SessionDep,
):
    member = await _get_member(staff_id, user.school_id, session)
    row = await session.scalar(
        select(StaffPromotion).where(
            StaffPromotion.id == prom_id,
            StaffPromotion.staff_member_id == member.id,
        )
    )
    if not row:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Promotion record not found")
    await session.delete(row)
    await session.commit()


# ── Permissions ───────────────────────────────────────────────────────────────

@router.get("/{staff_id}/permissions", response_model=StaffPermissionsResponse,
            dependencies=[require(Permission.MANAGE_USERS)])
async def get_staff_permissions(
    staff_id: UUID,
    user: CurrentUser,
    redis: RedisDep,
    session: SessionDep,
):
    member = await _get_member(staff_id, user.school_id, session)

    overrides = list(await session.scalars(
        select(StaffPermission).where(
            StaffPermission.staff_member_id == member.id,
            StaffPermission.school_id == user.school_id,
        )
    ))

    linked = await _linked_user(member, session)
    resolved = await resolve_all_permissions(linked, redis, session) if linked else {}

    # Load assigned roles with their role detail
    role_rows: list[UserRoleResponse] = []
    if linked:
        ur_rows = list(await session.scalars(
            select(UserRole).where(UserRole.user_id == linked.id).options(
                selectinload(UserRole.role)
            )
        ))
        role_rows = [
            UserRoleResponse(
                id=ur.id,
                role=RoleResponse(
                    id=ur.role.id,
                    name=ur.role.name,
                    code=ur.role.code,
                    is_system_template=ur.role.is_system_template,
                ),
                assigned_at=ur.assigned_at,
            )
            for ur in ur_rows
        ]

    return StaffPermissionsResponse(
        staff_member_id=member.id,
        roles=role_rows,
        permissions=resolved,
        overrides=[PermissionOverrideResponse.model_validate(o) for o in overrides],
    )


@router.post("/{staff_id}/permissions", response_model=PermissionOverrideResponse, status_code=201,
             dependencies=[require(Permission.MANAGE_USERS)])
async def set_permission_override(
    staff_id: UUID,
    body: PermissionOverrideCreate,
    user: CurrentUser,
    redis: RedisDep,
    session: SessionDep,
):
    member = await _get_member(staff_id, user.school_id, session)

    if body.permission_key not in ALL_PERMISSIONS:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Unknown permission: {body.permission_key}")

    existing = await session.scalar(
        select(StaffPermission).where(
            StaffPermission.staff_member_id == member.id,
            StaffPermission.school_id == user.school_id,
            StaffPermission.permission_key == body.permission_key,
        )
    )

    now = datetime.now(UTC)
    if existing:
        existing.granted = body.granted
        existing.note = body.note
        existing.granted_by = user.id
        existing.granted_at = now
        override = existing
    else:
        override = StaffPermission(
            staff_member_id=member.id,
            school_id=user.school_id,
            permission_key=body.permission_key,
            granted=body.granted,
            granted_by=user.id,
            granted_at=now,
            note=body.note,
        )
        session.add(override)

    await session.commit()
    await session.refresh(override)

    linked = await _linked_user(member, session)
    if linked:
        await invalidate_cache(linked.id, redis)

    return PermissionOverrideResponse.model_validate(override)


@router.delete("/{staff_id}/permissions/{perm_key}", status_code=204,
               dependencies=[require(Permission.MANAGE_USERS)])
async def remove_permission_override(
    staff_id: UUID,
    perm_key: str,
    user: CurrentUser,
    redis: RedisDep,
    session: SessionDep,
):
    member = await _get_member(staff_id, user.school_id, session)

    override = await session.scalar(
        select(StaffPermission).where(
            StaffPermission.staff_member_id == member.id,
            StaffPermission.school_id == user.school_id,
            StaffPermission.permission_key == perm_key,
        )
    )
    if not override:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Permission override not found")

    await session.delete(override)
    await session.commit()

    linked = await _linked_user(member, session)
    if linked:
        await invalidate_cache(linked.id, redis)


# ── Role assignment (user_role) ───────────────────────────────────────────────

@router.get("/{staff_id}/roles", response_model=list[UserRoleResponse],
            dependencies=[require(Permission.MANAGE_USERS)])
async def list_roles(staff_id: UUID, user: CurrentUser, session: SessionDep):
    member = await _get_member(staff_id, user.school_id, session)
    linked = await _linked_user(member, session)
    if not linked:
        return []
    rows = list(await session.scalars(
        select(UserRole).where(UserRole.user_id == linked.id)
        .options(selectinload(UserRole.role))
    ))
    return [
        UserRoleResponse(
            id=ur.id,
            role=RoleResponse(
                id=ur.role.id, name=ur.role.name,
                code=ur.role.code, is_system_template=ur.role.is_system_template,
            ),
            assigned_at=ur.assigned_at,
        )
        for ur in rows
    ]


@router.post("/{staff_id}/roles", response_model=UserRoleResponse, status_code=201,
             dependencies=[require(Permission.MANAGE_USERS)])
async def assign_role(
    staff_id: UUID,
    body: RoleAssignRequest,
    user: CurrentUser,
    redis: RedisDep,
    session: SessionDep,
):
    member = await _get_member(staff_id, user.school_id, session)
    linked = await _linked_user(member, session)
    if not linked:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "No user account linked to this staff member")

    role = await session.scalar(select(StaffPosition).where(StaffPosition.id == body.role_id))
    if not role:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Role not found")

    existing = await session.scalar(
        select(UserRole).where(UserRole.user_id == linked.id, UserRole.role_id == body.role_id)
    )
    if existing:
        raise HTTPException(status.HTTP_409_CONFLICT, "Role already assigned")

    ur = UserRole(
        user_id=linked.id,
        role_id=body.role_id,
        assigned_by=user.id,
        assigned_at=datetime.now(UTC),
    )
    session.add(ur)
    await session.commit()
    await session.refresh(ur)
    await invalidate_cache(linked.id, redis)

    return UserRoleResponse(
        id=ur.id,
        role=RoleResponse(
            id=role.id, name=role.name,
            code=role.code, is_system_template=role.is_system_template,
        ),
        assigned_at=ur.assigned_at,
    )


@router.delete("/{staff_id}/roles/{user_role_id}", status_code=204,
               dependencies=[require(Permission.MANAGE_USERS)])
async def remove_role(
    staff_id: UUID,
    user_role_id: UUID,
    user: CurrentUser,
    redis: RedisDep,
    session: SessionDep,
):
    member = await _get_member(staff_id, user.school_id, session)
    linked = await _linked_user(member, session)
    if not linked:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "No user account linked to this staff member")

    ur = await session.scalar(
        select(UserRole).where(UserRole.id == user_role_id, UserRole.user_id == linked.id)
    )
    if not ur:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Role assignment not found")

    await session.delete(ur)
    await session.commit()
    await invalidate_cache(linked.id, redis)


# ── Bulk upload ───────────────────────────────────────────────────────────────

BULK_REQUIRED = {"first_name", "last_name", "category"}

_TEMPLATE_COLUMNS = [
    ("first_name",              "First Name *",          18),
    ("middle_name",             "Middle Name",           18),
    ("last_name",               "Last Name *",           18),
    ("staff_id",                "School Staff ID",       16),
    ("category",                "Category *",            16),
    ("employment_type",         "Employment Type",       18),
    ("gender",                  "Gender",                12),
    ("date_of_birth",           "Date of Birth",         14),
    ("phone",                   "Phone",                 16),
    ("personal_email",          "Personal Email",        24),
    ("designation",             "Designation",           18),
    ("date_joined",             "Date Joined",           14),
    ("ges_staff_id",            "GES Staff ID",          16),
    ("ssnit_no",                "SSNIT No.",             16),
    ("registered_no",           "Registered No.",        16),
    ("licence_no",              "Licence No.",           16),
    ("address",                 "Home Address",          28),
    ("emergency_contact_name",  "Emergency Contact Name",22),
    ("emergency_contact_phone", "Emergency Contact Phone",22),
]

_DROPDOWNS: dict[str, list[str]] = {
    "category":        ["TEACHING", "NON-TEACHING"],
    "employment_type": ["PERMANENT", "CONTRACT", "VOLUNTEER", "GES_POSTED"],
    "gender":          ["MALE", "FEMALE", "OTHER"],
    "designation":     ["TEACHER", "HEADTEACHER", "ASSISTANT_HEAD", "BURSAR"],
}

_EXAMPLE_ROWS = [
    {
        "first_name": "Kwame", "middle_name": "Asante", "last_name": "Mensah",
        "staff_id": "STF001", "category": "TEACHING", "employment_type": "PERMANENT",
        "gender": "MALE", "date_of_birth": "1985-04-12", "phone": "0244123456",
        "personal_email": "k.mensah@school.edu.gh", "designation": "TEACHER",
        "date_joined": "2018-09-01", "ges_staff_id": "GES12345",
    },
    {
        "first_name": "Abena", "last_name": "Owusu",
        "staff_id": "STF002", "category": "NON-TEACHING", "employment_type": "PERMANENT",
        "gender": "FEMALE", "date_of_birth": "1990-07-20", "phone": "0201987654",
        "personal_email": "a.owusu@school.edu.gh", "designation": "BURSAR",
        "date_joined": "2020-01-15", "ssnit_no": "SSNIT12345",
    },
]


def _hex_no_hash(hex_color: str) -> str:
    return hex_color.lstrip("#").upper()

def _darken_hex(hex_color: str, factor: float = 0.6) -> str:
    h = _hex_no_hash(hex_color)
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return f"{int(r*factor):02X}{int(g*factor):02X}{int(b*factor):02X}"

def _tint_hex(hex_color: str, strength: float = 0.12) -> str:
    """Blend accent toward white. strength=0.12 → very pale tint."""
    h = _hex_no_hash(hex_color)
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    r2 = int(r * strength + 255 * (1 - strength))
    g2 = int(g * strength + 255 * (1 - strength))
    b2 = int(b * strength + 255 * (1 - strength))
    return f"{r2:02X}{g2:02X}{b2:02X}"

def _border_hex(hex_color: str, strength: float = 0.25) -> str:
    return _tint_hex(hex_color, strength)


@router.get("/bulk/template")
async def download_template(user: CurrentUser, session: SessionDep):
    """Return a formatted Excel template with dropdown validations."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    from app.models.school import School
    school = await session.scalar(select(School).where(School.id == user.school_id))
    school_name = school.name if school else "School"
    accent = school.accent_color if school else "#185FA5"

    # Derive palette from school accent color
    c_accent  = _hex_no_hash(accent)          # header row fill
    c_dark    = _darken_hex(accent, 0.55)     # title row fill
    c_tint    = _tint_hex(accent, 0.08)       # subtitle strip
    c_example = _tint_hex(accent, 0.12)       # example rows
    c_border  = _border_hex(accent, 0.30)     # cell border colour

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Staff Import"

    total_cols = len(_TEMPLATE_COLUMNS)

    # ── Title row ─────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1, value=f"{school_name} — Staff Import Template")
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=c_dark)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # ── Subtitle row ──────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    sub = ws.cell(row=2, column=1,
        value="Required columns: first_name, last_name, category  ·  Date format: YYYY-MM-DD  ·  Delete example rows before importing")
    sub.font = Font(italic=True, size=9, color="444444")
    sub.fill = PatternFill("solid", fgColor=c_tint)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Header row ────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor=c_accent)
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color=c_border)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (field, label, col_width) in enumerate(_TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[3].height = 30

    # ── Example rows ──────────────────────────────────────────────
    example_fill = PatternFill("solid", fgColor=c_example)
    fields = [f for f, _, _ in _TEMPLATE_COLUMNS]

    for row_offset, example in enumerate(_EXAMPLE_ROWS, start=4):
        for col_idx, field in enumerate(fields, start=1):
            val = example.get(field, "")
            cell = ws.cell(row=row_offset, column=col_idx, value=val)
            cell.fill = example_fill
            cell.font = Font(size=10, color="333333")
            cell.alignment = Alignment(vertical="center")
            cell.border = border
        ws.row_dimensions[row_offset].height = 18

    # ── Data rows (pre-format for up to 500 rows) ─────────────────
    data_font = Font(size=10)
    for row in range(6, 506):
        for col_idx in range(1, total_cols + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = data_font
            cell.alignment = Alignment(vertical="center")
            cell.border = border
        ws.row_dimensions[row].height = 18

    # ── Dropdown validations ──────────────────────────────────────
    for col_idx, (field, _label, _w) in enumerate(_TEMPLATE_COLUMNS, start=1):
        if field not in _DROPDOWNS:
            continue
        col_letter = get_column_letter(col_idx)
        formula = '"' + ",".join(_DROPDOWNS[field]) + '"'
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="Invalid value",
            error=f'Use the dropdown to select a valid {field.replace("_", " ")}.',
        )
        dv.sqref = f"{col_letter}4:{col_letter}505"
        ws.add_data_validation(dv)

    # ── Freeze panes below header ─────────────────────────────────
    ws.freeze_panes = "A4"

    # Serialize to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = school_name.replace(" ", "_").replace("/", "-")
    filename = f"{safe_name}_Staff_Import_Template.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
BULK_OPTIONAL = {
    "staff_id", "middle_name", "gender", "date_of_birth", "phone",
    "personal_email", "address", "emergency_contact_name", "emergency_contact_phone",
    "employment_type", "designation", "date_joined",
    "ges_staff_id", "registered_no", "licence_no", "ssnit_no",
}


def _parse_csv(content: bytes) -> list[dict]:
    import csv
    text = content.decode("utf-8-sig")
    # Strip comment lines and blank lines before feeding to DictReader
    # so templates with school metadata headers (# ...) parse cleanly
    data_lines = [l for l in text.splitlines() if l.strip() and not l.strip().startswith("#")]
    reader = csv.DictReader(io.StringIO("\n".join(data_lines)))
    return [row for row in reader]


def _parse_excel(content: bytes) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "Install openpyxl to support Excel uploads",
        )
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:]]


@router.post("/bulk", response_model=BulkUploadResponse, status_code=201,
             dependencies=[require(Permission.MANAGE_STAFF)])
async def bulk_upload(
    user: CurrentUser,
    session: SessionDep,
    file: UploadFile = File(...),
):

    content = await file.read()
    fname = (file.filename or "").lower()

    if fname.endswith(".csv"):
        rows = _parse_csv(content)
    elif fname.endswith((".xlsx", ".xls")):
        rows = _parse_excel(content)
    else:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Upload a .csv or .xlsx file")

    created = 0
    skipped = 0
    errors: list[BulkRowError] = []

    for i, raw in enumerate(rows, start=2):  # row 1 = header
        # Normalize keys
        row = {k.strip().lower(): (str(v).strip() if v is not None else "") for k, v in raw.items()}

        # Skip entirely blank rows
        if not any(row.values()):
            skipped += 1
            continue

        # Required field check
        missing = [f for f in BULK_REQUIRED if not row.get(f)]
        if missing:
            errors.append(BulkRowError(row=i, field=missing[0], message=f"Required: {missing[0]}"))
            continue

        # Build kwargs
        kwargs: dict = {
            "school_id": user.school_id,
            "first_name": row["first_name"],
            "last_name": row["last_name"],
            "category": row.get("category", "TEACHING").upper(),
        }
        for field in BULK_OPTIONAL:
            val = row.get(field)
            if val:
                if field in ("date_of_birth", "date_joined"):
                    try:
                        kwargs[field] = date.fromisoformat(val)
                    except ValueError:
                        errors.append(BulkRowError(row=i, field=field, message=f"Invalid date: {val}"))
                        continue
                else:
                    kwargs[field] = val

        if "middle_name" in row and row["middle_name"]:
            kwargs["middle_name"] = row["middle_name"]
        if "employment_type" in kwargs:
            kwargs["employment_type"] = kwargs["employment_type"].upper()

        try:
            member = StaffMember(**kwargs)
            session.add(member)
            await session.flush()
            created += 1
        except Exception as exc:
            errors.append(BulkRowError(row=i, field=None, message=str(exc)))

    if errors and created == 0:
        await session.rollback()
    else:
        await session.commit()

    return BulkUploadResponse(created=created, skipped=skipped, errors=errors)
