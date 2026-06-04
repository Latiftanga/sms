import io
from datetime import date
from uuid import UUID

from fastapi import APIRouter, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select

from app.api.deps import CurrentUser, SessionDep, require
from app.core.permissions import Permission
from app.models.school import School
from app.models.staff import StaffMember
from app.schemas.staff import BulkRowError, BulkUploadResponse

router = APIRouter()

BULK_REQUIRED = {"first_name", "last_name", "category"}
BULK_OPTIONAL = {
    "staff_id", "middle_name", "gender", "date_of_birth", "phone",
    "personal_email", "address", "emergency_contact_name", "emergency_contact_phone",
    "employment_type", "designation", "date_joined",
    "ges_staff_id", "registered_no", "licence_no", "ssnit_no",
}

_TEMPLATE_COLUMNS = [
    ("first_name",              "First Name *",          18),
    ("middle_name",             "Middle Name",           18),
    ("last_name",               "Last Name *",           18),
    ("staff_id",                "School Staff ID",       16),
    ("category",                "Category *",            16),
    ("employment_type",         "Employment Type",       18),
    ("gender",                  "Gender",                12),
    ("date_of_birth",           "Date of Birth",         14),
    ("phone",                   "Phone",                 16),
    ("personal_email",          "Personal Email",        24),
    ("designation",             "Designation",           18),
    ("date_joined",             "Date Joined",           14),
    ("ges_staff_id",            "GES Staff ID",          16),
    ("ssnit_no",                "SSNIT No.",             16),
    ("registered_no",           "Registered No.",        16),
    ("licence_no",              "Licence No.",           16),
    ("address",                 "Home Address",          28),
    ("emergency_contact_name",  "Emergency Contact Name",22),
    ("emergency_contact_phone", "Emergency Contact Phone",22),
]

_DROPDOWNS: dict[str, list[str]] = {
    "category":        ["TEACHING", "NON-TEACHING"],
    "employment_type": ["PERMANENT", "CONTRACT", "VOLUNTEER", "GES_POSTED"],
    "gender":          ["MALE", "FEMALE", "OTHER"],
    "designation":     ["TEACHER", "HEADTEACHER", "ASSISTANT_HEAD", "BURSAR", "HOUSEMASTER", "SENIOR_HOUSEMASTER"],
}

_EXAMPLE_ROWS = [
    {
        "first_name": "Kwame", "middle_name": "Asante", "last_name": "Mensah",
        "staff_id": "STF001", "category": "TEACHING", "employment_type": "PERMANENT",
        "gender": "MALE", "date_of_birth": "1985-04-12", "phone": "0244123456",
        "personal_email": "k.mensah@school.edu.gh", "designation": "TEACHER",
        "date_joined": "2018-09-01", "ges_staff_id": "GES12345",
    },
    {
        "first_name": "Abena", "last_name": "Owusu",
        "staff_id": "STF002", "category": "NON-TEACHING", "employment_type": "PERMANENT",
        "gender": "FEMALE", "date_of_birth": "1990-07-20", "phone": "0201987654",
        "personal_email": "a.owusu@school.edu.gh", "designation": "BURSAR",
        "date_joined": "2020-01-15", "ssnit_no": "SSNIT12345",
    },
]

_LABEL_TO_FIELD: dict[str, str] = {
    label.lower().rstrip(" *"): field
    for field, label, _ in _TEMPLATE_COLUMNS
}


def _hex_no_hash(hex_color: str) -> str:
    return hex_color.lstrip("#").upper()

def _darken_hex(hex_color: str, factor: float = 0.6) -> str:
    h = _hex_no_hash(hex_color)
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return f"{int(r*factor):02X}{int(g*factor):02X}{int(b*factor):02X}"

def _tint_hex(hex_color: str, strength: float = 0.12) -> str:
    h = _hex_no_hash(hex_color)
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    r2 = int(r * strength + 255 * (1 - strength))
    g2 = int(g * strength + 255 * (1 - strength))
    b2 = int(b * strength + 255 * (1 - strength))
    return f"{r2:02X}{g2:02X}{b2:02X}"

def _border_hex(hex_color: str, strength: float = 0.25) -> str:
    return _tint_hex(hex_color, strength)


@router.get("/bulk/template", dependencies=[require(Permission.MANAGE_STAFF)])
async def download_template(user: CurrentUser, session: SessionDep):
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    school = await session.scalar(select(School).where(School.id == user.school_id))
    school_name = school.name if school else "School"
    accent = school.accent_color if school else "#185FA5"

    c_accent  = _hex_no_hash(accent)
    c_dark    = _darken_hex(accent, 0.55)
    c_tint    = _tint_hex(accent, 0.08)
    c_example = _tint_hex(accent, 0.12)
    c_border  = _border_hex(accent, 0.30)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Staff Import"

    total_cols = len(_TEMPLATE_COLUMNS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1, value=f"{school_name} — Staff Import Template")
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=c_dark)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    sub = ws.cell(row=2, column=1,
        value="Required columns: first_name, last_name, category  ·  Date format: YYYY-MM-DD  ·  Delete example rows before importing")
    sub.font = Font(italic=True, size=9, color="444444")
    sub.fill = PatternFill("solid", fgColor=c_tint)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    header_fill = PatternFill("solid", fgColor=c_accent)
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color=c_border)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (field, label, col_width) in enumerate(_TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[3].height = 30

    example_fill = PatternFill("solid", fgColor=c_example)
    fields = [f for f, _, _ in _TEMPLATE_COLUMNS]

    for row_offset, example in enumerate(_EXAMPLE_ROWS, start=4):
        for col_idx, field in enumerate(fields, start=1):
            val = example.get(field, "")
            cell = ws.cell(row=row_offset, column=col_idx, value=val)
            cell.fill = example_fill
            cell.font = Font(size=10, color="333333")
            cell.alignment = Alignment(vertical="center")
            cell.border = border
        ws.row_dimensions[row_offset].height = 18

    data_font = Font(size=10)
    for row in range(6, 506):
        for col_idx in range(1, total_cols + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = data_font
            cell.alignment = Alignment(vertical="center")
            cell.border = border
        ws.row_dimensions[row].height = 18

    for col_idx, (field, _label, _w) in enumerate(_TEMPLATE_COLUMNS, start=1):
        if field not in _DROPDOWNS:
            continue
        col_letter = get_column_letter(col_idx)
        formula = '"' + ",".join(_DROPDOWNS[field]) + '"'
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="Invalid value",
            error=f'Use the dropdown to select a valid {field.replace("_", " ")}.',
        )
        dv.sqref = f"{col_letter}4:{col_letter}505"
        ws.add_data_validation(dv)

    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = school_name.replace(" ", "_").replace("/", "-")
    filename = f"{safe_name}_Staff_Import_Template.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _parse_csv(content: bytes) -> list[dict]:
    import csv
    text = content.decode("utf-8-sig")
    data_lines = [l for l in text.splitlines() if l.strip() and not l.strip().startswith("#")]
    reader = csv.DictReader(io.StringIO("\n".join(data_lines)))
    return [row for row in reader]


def _parse_excel(content: bytes) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "Install openpyxl to support Excel uploads",
        )
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    known = set(_LABEL_TO_FIELD) | {f for f, _, _ in _TEMPLATE_COLUMNS}
    header_idx = 0
    for i, row in enumerate(rows):
        first = str(row[0]).strip().lower().rstrip(" *") if row[0] is not None else ""
        if first in known:
            header_idx = i
            break

    raw_headers = [str(h).strip().lower().rstrip(" *") if h is not None else "" for h in rows[header_idx]]
    headers = [_LABEL_TO_FIELD.get(h, h) for h in raw_headers]

    return [
        dict(zip(headers, row))
        for row in rows[header_idx + 1:]
    ]


@router.post("/bulk", response_model=BulkUploadResponse, status_code=201,
             dependencies=[require(Permission.MANAGE_STAFF)])
async def bulk_upload(
    user: CurrentUser,
    session: SessionDep,
    file: UploadFile = File(...),
):
    content = await file.read()
    fname = (file.filename or "").lower()

    if fname.endswith(".csv"):
        rows = _parse_csv(content)
    elif fname.endswith((".xlsx", ".xls")):
        rows = _parse_excel(content)
    else:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Upload a .csv or .xlsx file")

    created = 0
    skipped = 0
    errors: list[BulkRowError] = []

    for i, raw in enumerate(rows, start=2):
        row = {k.strip().lower(): (str(v).strip() if v is not None else "") for k, v in raw.items()}

        if not any(row.values()):
            skipped += 1
            continue

        missing = [f for f in BULK_REQUIRED if not row.get(f)]
        if missing:
            errors.append(BulkRowError(row=i, field=missing[0], message=f"Required: {missing[0]}"))
            continue

        kwargs: dict = {
            "school_id": user.school_id,
            "first_name": row["first_name"],
            "last_name": row["last_name"],
            "category": row.get("category", "TEACHING").upper(),
        }
        for field in BULK_OPTIONAL:
            val = row.get(field)
            if val:
                if field in ("date_of_birth", "date_joined"):
                    try:
                        kwargs[field] = date.fromisoformat(val)
                    except ValueError:
                        errors.append(BulkRowError(row=i, field=field, message=f"Invalid date: {val}"))
                        continue
                else:
                    kwargs[field] = val

        if "middle_name" in row and row["middle_name"]:
            kwargs["middle_name"] = row["middle_name"]
        if "employment_type" in kwargs:
            kwargs["employment_type"] = kwargs["employment_type"].upper()

        try:
            member = StaffMember(**kwargs)
            session.add(member)
            await session.flush()
            created += 1
        except Exception as exc:
            errors.append(BulkRowError(row=i, field=None, message=str(exc)))

    if errors and created == 0:
        await session.rollback()
    else:
        await session.commit()

    return BulkUploadResponse(created=created, skipped=skipped, errors=errors)
